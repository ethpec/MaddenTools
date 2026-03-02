# Imports
import pandas as pd
import random
from openpyxl.styles import PatternFill, Alignment

# ==============================
# File Paths
# ==============================

file_path = 'Files/Madden26/IE/Season2/Player.xlsx'
scouting_file_path = 'Files/Madden26/IE/Season2/AllScoutInfo.xlsx'

# ==============================
# Load Player DataFrame
# ==============================

df = pd.read_excel(file_path)

# ==============================
# Convert numeric columns to object
# ==============================

# This makes it safe to assign letter grades
numeric_columns = df.select_dtypes(include='number').columns
df[numeric_columns] = df[numeric_columns].astype(object)

# ==============================
# Load other sheets
# ==============================

positions_df = pd.read_excel(scouting_file_path, sheet_name="Positions")
spline_df = pd.read_excel(scouting_file_path, sheet_name="spline")
sheet_1680_df = pd.read_excel(scouting_file_path, sheet_name="1680")

# Track original BEFORE edits
original_df = df.copy()

# Set Positions index (this one we know)
positions_df.set_index("Position", inplace=True)

position_groups = {
    "QB": ["QB"],
    "RB": ["HB", "RB"],
    "WR": ["WR"],
    "TE": ["TE"],
    "OL": ["LT", "LG", "C", "RG", "RT"],
    "DL": ["LE", "RE", "DT"],
    "LB": ["MLB", "LOLB", "ROLB"],
    "CB": ["CB"],
    "SAF": ["FS", "SS"],
    "ST": ["K", "P"]
}

# ==============================
# Process Draft Players
# ==============================

for index, player in df[df["ContractStatus"] == "Draft"].iterrows():

    position = player["Position"]

    if position not in positions_df.index:
        continue

    pos_row = positions_df.loc[position]

    for attribute, spline_key in pos_row.items():

        if spline_key == 0:
            continue

        if attribute not in df.columns:
            continue

        # spline_key is the row index
        if spline_key >= len(spline_df):
            print(f"spline_key {spline_key} out of spline_df range")
            continue

        spline_row = spline_df.iloc[int(spline_key) - 2]
        x_row_value = spline_row["x row"]

        # Use x_row_value as row index
        if x_row_value >= len(sheet_1680_df):
            continue

        int_row = sheet_1680_df.iloc[int(x_row_value) - 2]

        int1 = int_row["int1"]
        int2 = int_row["int2"]
        int3 = int_row["int3"]
        int4 = int_row["int4"]
        int5 = int_row["int5"]
        int6 = int_row["int6"]

        player_value = player[attribute]

        # ==============================
        # Letter Grade Logic
        # ==============================
        if int5 == 0:
            if player_value < int1:
                df.at[index, attribute] = "F"

            elif player_value < int2:
                df.at[index, attribute] = "D"

            elif player_value < int3:
                df.at[index, attribute] = "C"

            elif player_value < int4:
                df.at[index, attribute] = "B"

            elif player_value >= int4:
                df.at[index, attribute] = "A"


        else:
            if player_value < int1:
                df.at[index, attribute] = "1-Poor"

            elif player_value < int2:
                df.at[index, attribute] = "2-Marginal"

            elif player_value < int3:
                df.at[index, attribute] = "3-Decent"

            elif player_value < int4:
                df.at[index, attribute] = "4-Solid"

            elif player_value < int5:
                df.at[index, attribute] = "5-Good"

            elif player_value < int6:
                df.at[index, attribute] = "6-Great"

            else:
                df.at[index, attribute] = "7-Elite"

# ==============================
# Keep only changed columns
# ==============================

columns_to_keep = [
    column for column in df.columns
    if not df[column].equals(original_df[column])
]

# Ensure essential columns are at front
front_columns = ["ContractStatus", "PLYR_DRAFTROUND", "PLYR_DRAFTPICK", "FirstName", "LastName", "Height", "Weight", "Age", "Position"]

# Only add front_columns if they exist in the DataFrame
front_columns = [col for col in front_columns if col in df.columns]

# Combine front columns with changed columns, avoiding duplicates
columns_to_keep = front_columns + [col for col in columns_to_keep if col not in front_columns]

# Reorder the DataFrame
df = df[columns_to_keep]

# ==============================
# Column Cleanup (Export Formatting Only)
# ==============================

def clean_column_name(col):

    # First remove common suffixes
    col = col.replace("Rating", "").replace("Grade", "")

    # Custom abbreviations
    abbreviations = {

        # Awareness
        "Awareness": "AWR",

        # QB
        "ThrowPower": "THP",
        "ThrowUnderPressure": "TUP",
        "ThrowOnTheRun": "TOR",
        "ThrowAccuracyShort": "SAC",
        "ThrowAccuracyMid": "MAC",
        "ThrowAccuracyDeep": "DAC",
        "PlayAction": "PAC",
        "BreakSack": "BSK",

        # Ball Carrier
        "Speed": "SPD",
        "Acceleration": "ACC",
        "Agility": "AGI",
        "ChangeOfDirection": "COD",
        "Carrying": "CAR",
        "BreakTackle": "BTK",
        "Trucking": "TRK",
        "StiffArm": "SFA",
        "JukeMove": "JKM",
        "SpinMove": "SPM",
        "BCVision": "BCV",

        # Receiving
        "Catching": "CTH",
        "CatchInTraffic": "CIT",
        "SpectacularCatch": "SPC",
        "Release": "RLS",
        "ShortRouteRunning": "SRR",
        "MediumRouteRunning": "MRR",
        "DeepRouteRunning": "DRR",

        # Blocking
        "PassBlock": "PBK",
        "PassBlockPower": "PBP",
        "PassBlockFinesse": "PBF",
        "RunBlock": "RBK",
        "RunBlockPower": "RBP",
        "RunBlockFinesse": "RBF",
        "ImpactBlocking": "IBL",
        "LeadBlock": "LBK",

        # Defensive
        "ManCoverage": "MCV",
        "ZoneCoverage": "ZCV",
        "Press": "PRS",
        "Pursuit": "PUR",
        "Tackle": "TAK",
        "HitPower": "HIT",
        "PowerMoves": "PMV",
        "FinesseMoves": "FMV",
        "PlayRecognition": "PRC",
        "BlockShedding": "BSH",

        # Physical
        "Strength": "STR",
        "Jumping": "JMP",
        "Stamina": "STA",
        "Toughness": "TGH",
        "Injury": "INJ",

        # Special Teams
        "KickPower": "KPW",
        "KickAccuracy": "KAC",
        "KickReturn": "KR",
        "LongSnap": "LS"
    }

    if col in abbreviations:
        return abbreviations[col]

    # Special cases
    if col == "PLYR_DRAFTROUND":
        return "Round"
    if col == "PLYR_DRAFTPICK":
        return "Pick"

    return col

# ==============================
# Grade Color Fills
# ==============================

GRADE_FILLS = {
    "A":          PatternFill("solid", fgColor="81c784"),
    "B":          PatternFill("solid", fgColor="c8e6c9"),
    "C":          PatternFill("solid", fgColor="fff176"),
    "D":          PatternFill("solid", fgColor="ffcc80"),
    "F":          PatternFill("solid", fgColor="ef9a9a"),
    "7-Elite":    PatternFill("solid", fgColor="81c784"),
    "6-Great":    PatternFill("solid", fgColor="a5d6a7"),
    "5-Good":     PatternFill("solid", fgColor="c8e6c9"),
    "4-Solid":    PatternFill("solid", fgColor="fff9c4"),
    "3-Decent":   PatternFill("solid", fgColor="fff176"),
    "2-Marginal": PatternFill("solid", fgColor="ffcc80"),
    "1-Poor":     PatternFill("solid", fgColor="ef9a9a"),
}

LEFT_ALIGN = Alignment(horizontal="left")

def apply_grade_colors(ws):
    ws.auto_filter.ref = ws.dimensions

    # Track max content length per column for auto-width
    col_widths = {}

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = LEFT_ALIGN
            col_letter = cell.column_letter
            cell_len = len(str(cell.value)) if cell.value is not None else 0
            if cell.row == 1:
                cell_len += 2  # extra room for filter dropdown arrow
            col_widths[col_letter] = max(col_widths.get(col_letter, 0), cell_len)

            if cell.row > 1:
                fill = GRADE_FILLS.get(str(cell.value) if cell.value is not None else "")
                if fill:
                    cell.fill = fill

    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width + 2

# ==============================
# Save Output (Multi-Sheet by Position Group + All)
# ==============================

output_filename = 'Files/Madden26/IE/Season2/Draft_LetterGrades.xlsx'

# Base columns always included
base_columns = ["PLYR_DRAFTROUND", "PLYR_DRAFTPICK", "FirstName", "LastName", "Height", "Weight", "Age", "Position"]

# Keep only Draft players for export
draft_df = df[df["ContractStatus"] == "Draft"].copy()

# Add Rank: overall pick number across all rounds
draft_df["Rank"] = draft_df["PLYR_DRAFTPICK"].astype(int) + 32 * (draft_df["PLYR_DRAFTROUND"].astype(int) - 1)

# Convert Weight to true value
draft_df["Weight"] = draft_df["Weight"].astype(int) + 160

# Convert Height from inches to feet'inches" format
draft_df["Height"] = draft_df["Height"].astype(int).apply(lambda h: f"{h // 12}'{h % 12}\"")

# Sort draft_df for the "All" sheet
draft_df.sort_values(by=["PLYR_DRAFTROUND", "PLYR_DRAFTPICK"], inplace=True)

with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:

    # Rename map for base column cleanup
    base_rename_map = {
        "PLYR_DRAFTROUND": "Round",
        "PLYR_DRAFTPICK": "Pick"
    }

    # Save All draft players sheet first
    draft_df[["Rank"] + base_columns].rename(columns=base_rename_map).to_excel(
        writer,
        sheet_name="All",
        index=False
    )
    apply_grade_colors(writer.sheets["All"])

    # Preferred column ordering per sheet (using abbreviations)
    position_column_order = {
        "QB":  ["THP", "SAC", "MAC", "DAC", "TUP", "TOR", "PAC", "BSK", "SPD", "ACC", "AGI", "STR", "AWR"],
        "RB":  ["SPD", "ACC", "AGI", "COD", "CAR", "BTK", "TRK", "SFA", "JKM", "SPM", "BCV", "STR", "SRR", "CTH", "CIT"],
        "WR":  ["SRR", "MRR", "DRR", "RLS", "CTH", "CIT", "SPC", "SPD", "ACC", "AGI", "STR"],
        "TE":  ["CTH", "CIT", "SPC", "RLS", "SRR", "MRR", "DRR", "PBK", "RBK", "IBL", "SPD", "ACC", "AGI", "STR", "BTK", "TRK"],
        "OL":  ["PBK", "PBP", "PBF", "RBK", "RBP", "RBF", "IBL", "LBK", "STR", "AGI", "ACC"],
        "DL":  ["PMV", "FMV", "BSH", "TAK", "HIT", "PUR", "PRC", "STR", "SPD", "ACC", "AGI"],
        "LB":  ["TAK", "HIT", "PUR", "MCV", "ZCV", "PRC", "PMV", "FMV", "BSH", "SPD", "ACC", "AGI", "STR"],
        "CB":  ["MCV", "ZCV", "PRS", "PRC", "TAK", "SPD", "ACC", "AGI", "STR"],
        "SAF": ["MCV", "ZCV", "PRC", "TAK", "HIT", "PUR", "BSH", "SPD", "ACC", "AGI", "STR"],
        "ST":  ["KPW", "KAC", "KR", "LS", "SPD", "ACC"],
    }

    for sheet_name, positions in position_groups.items():

        group_df = draft_df[draft_df["Position"].isin(positions)].copy()

        if group_df.empty:
            continue

        # Find which columns were actually changed for THIS position group
        changed_columns = [
            col for col in group_df.columns
            if col not in base_columns and col != "Rank"
            and not group_df[col].equals(
                original_df.loc[group_df.index, col]
            )
        ]

        # Apply preferred column order for this sheet if defined
        preferred_abbrevs = position_column_order.get(sheet_name, [])
        if preferred_abbrevs:
            abbrev_to_col = {clean_column_name(col): col for col in changed_columns}
            ordered = [abbrev_to_col[a] for a in preferred_abbrevs if a in abbrev_to_col]
            remaining = [col for col in changed_columns if col not in ordered]
            changed_columns = ordered + remaining

        # Final columns for this sheet
        export_columns = ["Rank"] + base_columns + changed_columns

        # Sort by Draft Round and Draft Pick
        group_df.sort_values(by=["PLYR_DRAFTROUND", "PLYR_DRAFTPICK"], inplace=True)

        rename_map = {col: clean_column_name(col) for col in export_columns}

        group_df[export_columns].rename(columns=rename_map).to_excel(
            writer,
            sheet_name=sheet_name,
            index=False
        )
        apply_grade_colors(writer.sheets[sheet_name])

print("Draft letter grades generated successfully.")